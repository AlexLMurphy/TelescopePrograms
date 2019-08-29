""" 
AnalyzeCRData.py

Author : A. Murphy
Email  : almurp20@g.holycross.edu
Updated: 7/15/2019

This program is inteded to be used to analyze data collected by the array of
cosmic-ray detection telescopes at the College of the Holy Cross in Worcester, MA.
This array of telescopes saves a hexadecimal timestamp to a file for each arrival
of coincident photomultiplier pulses. Thus, this program was created to convert
this data to sensible units of seconds and to use coincidence gating to confirm
that a muon shower hit multiple telescopes in the array. Coincidence gating allows
us to distinguish muons detected by the telescope array from other sources of noise.
We correct for small timeskips in the seconds while converting from hex, but avoid
overcorrecting the data.

A program with 3 main functions:
1) Concatenate several files of hexadecimal timestamps of cosmic ray telescope data
This requires several files whose names are F*.txt, whose contents
This program saves the contents of those files into one (1) large file chosen by the user,
and orders those contents according to their original file names. The other functions of
this program further sort and correct the timestamps.

Then, convert the combined HEX file (created from function 1) to decimal format. This function
corrects time skips in the data, starts from when the GPS was reset in the file (if applicable!),
and saves the decimal data to a new output file in two columns. The first column is seconds
counted by the GPS, and the second column is subseconds (already divided by a factor of 250
as a property of our electronic board IC LM555).

Lastly, an excel file is created for the light curve of that decimal data.

2) Do coincidence-scanning for multiple telescopes. Coincident events are saved to an output
file. The coincident events are designated by the first recorded time within the coincidence.
The first number in the coincident events file is the amount of telescopes that were coincident.

3) Convert decimal data back to formatted hexadecimal data. You should only have to use
this function if you have accidentally deleted a hex file or cannot find it. 

Recent changes:
Added several user input instances to streamline the application without
having to modify constants or commenting.
Added GUI interfaces so that it is easier to interact with the program.
Added sample statistics to the automatically generated light curve,
and these sample statistics filter out bins that are less than 1/4 of the mean
Fixed GPS string parsing to work with strings that have certain missing characters.
Added interval statistics analysis to the light curve function.

TODO:
Fix how std deviation is calculated in createLightCurve. It cannot use SUMIF.
"""

import os, os.path # For changing the directory
import glob # File operations within a directory
import math # Specifically for converting hex back to decimal with ceiling function
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from datetime import date # Just to suggest the current day for a file name
import PySimpleGUI as sg # GUI library

# Default directory can be changed for convenience.
directory = "F:\\" # Second backslash required to keep string format
defaultDirectory = os.getcwd()
namePattern = "F*.txt"
LM555FACTOR = 255 # This number changes if the users says the Arduino was used
ArduinoDeadTime = 0.001 # In seconds. Update this number if the dead time changes, or if new measurements are made. 
BS2DeadTime = 0.275 # Also in seconds.

def readfile(filename):
    with open(filename, "r") as file:                # open file
        data = []                                    # declare list
        for entry in file:                           # for each value in the file...
             data.append(entry.rstrip()) # append to data array, rstrip removes appended whitespace
    file.close()
    return data

def writefile(datalist, filename):
    fp = open(filename, "a")
    for entry in datalist: 
        print(entry, file=fp)
    fp.close()

# Requires a gps string with an 8-digit HEX timestamp at the end.
# Todo: Fix so that it returns 0 whenever the GPS string cannot be processed.
def calc_starting_time(gpsString):
    gpsdata = []
    if ("." in gpsString) and ("," in gpsString) and len(gpsString) >= 12 and gpsString[12] == "A":
        # Note: we should avoid using gpsString[12] and instead use an element in tempList below.
        pass
    else:
        return 0 # Function cannot process this GPS string.
     # It was a GPS string and we need to parse it!
    tempList = gpsString.split(",")
    # Parse the GPS string.
    # Note that sometimes the GPS string is not fully intact. Usually, it is missing characters.
    # To try parsing the string, we assume every comma is present. This way, it is harder for
    # missing characters to interfere.
    # The "straightforward" way would be to use date = gpsString[49:53], but this assumes all 48
    # characters before are present in the string. 
    timeStamp = tempList[1]
    timeStamp = timeStamp.split(".")[0] # Remove the decimal seconds from the timeStamp.
                                        # they were always ".000" to begin with.
    date = tempList[9]
    date = date[0:2] + date[2:4]
    
    currTime = gpsString[-8:].rstrip() # Still a string - rstrip() is used because the end of the
                                 # GPS string has a newline character.
    try:
        currTime = int(currTime[:-2], 16) # + int(currTime[-2:])/255 ignores subseconds for now
    except ValueError:
        # GPS string was missing the clock time at the end. But this could only happen
        # if an old version of the Arduino code was uploaded.
        return "Error"
    
    if "," not in (timeStamp + date) and "." not in (timeStamp + date):
        # If both parts of the GPS string are good, we save it along with
        # the previous timestamp for time analysis later.
        gpsdata = [timeStamp, date, currTime, gpsString]
        # Note that gpsdata[2] is the decimal seconds of the timestamp before this GPS string

    else: # Actual error in the hex data from sloppy clock signals.
        return "GPS string error"       

    # Perform date calculations with the GPS data.
    # day arrays for fast month checking
    day31List = [1, 3, 5, 7, 8, 10, 12]
    day30List = [4, 6, 9, 11]
    timeStamps = []
    # The date part of the gps data has already been rearranged into "mmdd"
    # It's too early to convert it into an integer for underflow reasons.
    # It's reasonable to assume that the months start counting from 1, and
    # the days start counting from 1.
    date = gpsdata[1]
    months = int(date[2:4])
    days = int(date[0:2])
    
    tempString = gpsdata[0]
    # Calculate the time of the day, in seconds, from the gps string
    gpsTime = (int(tempString[0:2])*3600) + (int(tempString[2:4])*60) + (int(tempString[4:6]))

    # Now compare the gps time of day to the seconds that are being counted by the circuit
    # clocks. For example, if gps time is 18:00:00 = 18*3600seconds = 64800seconds, and
    # the clock's time is 1000seconds for the last detection, then we will say that the
    # data collection started at:
    # 64800seconds - 1000seconds = 63800seconds = 17hrs, 43minutes, 20seconds
    # Note that the time of day will NOT be accurate if the GPS time is never reset on the
    # circuit board.

    # How do we handle the GPS time passing midnight?
    # If a negative "startTime" occurs, then we can just subtract from the gps date.
    
    # How do we handle the seconds time surpassing 24 hours?
    # Modulate the seconds by 24hours = 86400seconds later
    startTime = gpsTime - gpsdata[2]
    
    # By the way, this start time becomes negative as soon as the GPS rolls over.
    subtractDays = startTime / 86400
    if subtractDays < 0:
        subtractDays = abs(int(subtractDays)) + 1
        days -= subtractDays
    while days <= 0:
        # Note: negative days must be handled twice, because negative days cannot be
        # used in the UTC conversion
        months -= 1
        if months < 1:
            months += 12 # Only executes to go from months=0 (null) to months=12 (december)
        if months in day31List:
            days += 31
        elif months in day30List:
            days += 30
        else: # February: Not checking for leap years sorry
            days += 28       
    
    startTime = abs(startTime % 86400) # Modulate the result after computing the amount of days
                                       # that should be subtracted 
    # This should work?     
    hours = int(startTime / 3600)
    startTime -= hours * 3600
    minutes = int(startTime  / 60)
    startTime -= minutes * 60
    seconds = int(startTime)

    if (months >= 1) and (months <= 12):
        # Lastly convert from UTC to Eastern time.
        # EDT (spring-summer) is UTC-4. EST (fall-winter) is UTC-5
        # Exact dates for daylight savings is Mar 10 and Nov 3

        # First, check through all of the easy months.
        if months >= 4 and months <= 10: # EDT
            hours -= 4
        elif months == 12 or months == 1 or months == 2: # EST
            hours -= 5
        elif months == 3: # March.
            if days <= 9: # EST
                hours -= 5
            else: # EDT
                hours -=4
        elif months == 11: # November.
            if days <= 3: # EDT
                hours -= 4
            else: # EST
                hours -=5

        # Handle negative (or 0) days at the end, again
        if days <= 0:
            months -= 1
            if months in day31List:
                days = 31 - abs(days)
            elif months in day30List:
                days = 30 - abs(days) 
            else: # February: Not checking for leap years sorry
                days = 28 - abs(days) 

        # Do some data conversion and then append the gps string. We use zfill to get the leading zeros back.
        entry = [str(hours).zfill(2), str(minutes).zfill(2), str(seconds).zfill(2), str(months).zfill(2), str(days).zfill(2)]
        return entry
    else:
        return 0 # months error somehow.
    
# New: Incomplete
# For every file conversion (the files are demarked by the GPS strings), use the GPS string
# header to resync the data.
# We have no way to know if a sync error occurred within a file if we use this method.
# We are fine with this because we know sync errors happen about once a day.
# Don't bother checking to see if the starting times are reasonable. Here's why:
# When two starting times don't match up, the first file (that was created earlier) gets
# deleted. This is because we know that the rollover or GPS time error occurred in the first
# file. Therefore, there is no point in checking to see if every file is "reasonable."
# We essentially only keep a file if the next file has the same starting time. 
def readHEXtoDEC_GPSSYNC(filename):
    global LM555FACTOR
    data = [] # holds decimal-converted regular timestamps.
    selectGpsTimes = [] # for use with GUI selection.
    start_time_array = [] # military time string array like: [hours, minutes, seconds, month, day]
    with open(filename, "r") as fp:
        for val in fp:
            try:
                int(val, 16)
            except ValueError:
                z = calc_starting_time(val)
##                print(val, z)
                start_time_array.append(z)
    fp.close()

    # Before we do anything, iterate through the start_time_array and find
    # the most frequent starting time. This is what will be used as the absolute reference
    # for calculating all of the timestamps in decimal seconds.
    # We refuse to use EST timestamps because this makes graphing difficult.
    # This decision will also mean that it's possible for some detections to have a negative timestamp.
    
    # Now we have an array of timestamps. Use GUI to ask the user for when the telescopes were started.
    layout = [[sg.Text("Look at these suggested start times from the GPS strings.")],
              [sg.Text("Pick the one(s) that makes the most sense.")],
              [sg.Text("Or just select nothing if they are all bad.")]]
    layout2 = []

    # Sort through the timestamps to get only unique ones? Potentially better to just pick
    # the most frequent timestamp and assume that's when the data collection started.
    uniqueTimes = [] # Holds the actual unique timestamps
    uniqueTimesCounts = [] # Holds the count of each unique timestamp
    seenValues = set()
    for item in start_time_array:
        t = tuple(item)
        if t not in seenValues:
            uniqueTimes.append(item)
            uniqueTimesCounts.append(int(1))
            seenValues.add(t)
        else:
            index = uniqueTimes.index(item)
            uniqueTimesCounts[index] += 1
            
    joined = []
    for x in range(len(uniqueTimes)):
        joined.append([uniqueTimes[x], uniqueTimesCounts[x]])

    joined = sorted(joined, key=lambda x: x[1], reverse=True)
    # The second "for" loop isn't really necessary, but it helps make the code more vertical
    # and not have too many characters on the same line
    formattedTimes = []
    amount = 0
    for values in joined:
        if amount < 12:
            theString = "Telescope recording began at (military time): "
            for x in range(len(values[0])):
                if x < 2:
                    theString += values[0][x]
                    theString += ":"
                elif x == 2:
                    theString += values[0][x]
                    theString += " and the date was: "
                elif x == 3:
                    theString += values[0][x]
                    theString += "/"
                else: # x >= 4
                    theString += values[0][x]

            theString += " (" + str(values[1]) + ")"       
            formattedTimes.append(theString) # append before concatenation
            layout2.append([sg.Checkbox(theString)])
            amount += 1
        else:
            print("ERROR: Too many different GPS strings for the program to display at once.\n")

    layout.append([sg.Frame("Possible telescope start times:", layout2, title_color="blue")])
    layout.append([sg.Submit()])
    window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
    event, values = window.Read()
    window.Close()

    for x in range(len(values)):
        if values[x] == True:
            selectGpsTimes.append(formattedTimes[x])
            
    startReferenceTime = joined[0][0] # This is a list of 5 strings: ["hh", "mm", "ss", "mm", dd"]
    
    return data, selectGpsTimes

def readHEXtoDEC(filename):
    global LM555FACTOR
    data = []                       # return array
    selectGpsTimes = []             # return array
    gpsdata = []                    # Array for analyzing the Arduino gps strings
    with open(filename, "r") as fp:
        lineNum = 1                     # line number counter for errors

        for val in fp:                  # for each value in the file...
            sec_string = val[:6]        # extract the seconds
            subsec_string = val[6:]     # extract sub-second time
            try:                
                sec = int(sec_string,16)    # convert hex string to integer
                subsec = int(subsec_string,16)
                entry = [sec, subsec]    # create integer timestamp
            except ValueError:
                if ("." in val) or ("," in val): # It was a GPS string and we need to parse it!
                    tempList = val.split(",")
                    if len(val) >= 12 and val[12] == "A":
                        # Parse the GPS string.
                        # Note that sometimes the GPS string is not fully intact. Usually, it is missing characters.
                        # To try parsing the string, we assume every comma is present. This way, it is harder for
                        # missing characters to interfere.
                        # The "straightforward" way would be to use date = val[49:53], but this assumes all 48
                        # characters before are present in the string. 
                        timeStamp = tempList[1]
                        timeStamp = timeStamp.split(".")[0] # Remove the decimal seconds from the timeStamp.
                                                            # they were always ".000" to begin with.
                        date = tempList[9]
                        date = date[0:2] + date[2:4]
                        
                        currTime = val[-8:].rstrip() # Still a string - rstrip() is used because the end of the
                                                     # GPS string has a newline character.
                        try:
                            currTime = int(currTime[:-2], 16) # + int(currTime[-2:])/255 ignores subseconds for now
                        except ValueError:
                            # GPS string was missing the clock time at the end. But this could only happen
                            # if an old version of the Arduino code was uploaded.
                            continue
                        if "," not in (timeStamp + date) and "." not in (timeStamp + date):
                            # If both parts of the GPS string are good, we save it along with
                            # the previous timestamp for time analysis later.
                            gpsdata.append([timeStamp, date, currTime, val])
                            # Note that gpsdata[2] is the decimal seconds of the timestamp before this GPS string

                else: # Actual error in the hex data from sloppy clock signals.
                    layout = [[sg.Text("There was a data conversion error at line number:", text_color="red"),
                               sg.Input(default_text=lineNum, disabled=True)],
                              [sg.Text("in the combined HEX file. This is because the data was bad.", text_color="red")],
                              [sg.Text("This is what the data looked like on that line:", text_color="red"),
                               sg.Input(default_text=val, disabled=True)],
                              [sg.Text("You should manually fix it so that it is a 8-digit HEX number", text_color="red")],
                              [sg.Ok()]]
                    window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
                    event, values = window.Read()
                    window.Close()
                    ValueErrorNum += 1       
                continue
            
            data.append(entry)          # append to data array
            lineNum += 1
    fp.close()

    # Process GPS strings if they were found in the file.
    if len(gpsdata) > 0:
        # day arrays for fast month checking
        day31List = [1, 3, 5, 7, 8, 10, 12]
        day30List = [4, 6, 9, 11]
        timeStamps = []
        for x in range(len(gpsdata)):
            # The date part of the gps data has already been rearranged into "mmdd"
            # It's too early to convert it into an integer for underflow reasons.
            # It's reasonable to assume that the months start counting from 1, and
            # the days start counting from 1.
            date = gpsdata[x][1]
            months = int(date[2:4])
            days = int(date[0:2])
            
            tempString = gpsdata[x][0]
            # Calculate the time of the day, in seconds, from the gps string
            gpsTime = (int(tempString[0:2])*3600) + (int(tempString[2:4])*60) + (int(tempString[4:6]))

            # Now compare the gps time of day to the seconds that are being counted by the circuit
            # clocks. For example, if gps time is 18:00:00 = 18*3600seconds = 64800seconds, and
            # the clock's time is 1000seconds for the last detection, then we will say that the
            # data collection started at:
            # 64800seconds - 1000seconds = 63800seconds = 17hrs, 43minutes, 20seconds
            # Note that the time of day will NOT be accurate if the GPS time is never reset on the
            # circuit board.

            # How do we handle the GPS time passing midnight?
            # If a negative "startTime" occurs, then we can just subtract from the gps date.
            
            # How do we handle the seconds time surpassing 24 hours?
            # Modulate the seconds by 24hours = 86400seconds later
            startTime = gpsTime - gpsdata[x][2]
            
            # By the way, this start time becomes negative as soon as the GPS rolls over.
            subtractDays = startTime / 86400
            if subtractDays < 0:
                subtractDays = abs(int(subtractDays)) + 1
                days -= subtractDays
            while days <= 0:
                # Note: negative days must be handled twice, because negative days cannot be
                # used in the UTC conversion
                months -= 1
                if months < 1:
                    months += 12 # Only executes to go from months=0 (null) to months=12 (december)
                if months in day31List:
                    days += 31
                elif months in day30List:
                    days += 30
                else: # February: Not checking for leap years sorry
                    days += 28       
            
            startTime = abs(startTime % 86400) # Modulate the result after computing the amount of days
                                               # that should be subtracted 
            # This should work?     
            hours = int(startTime / 3600)
            startTime -= hours * 3600
            minutes = int(startTime  / 60)
            startTime -= minutes * 60
            seconds = int(startTime)

            if (months >= 1) and (months <= 12):
                # Lastly convert from UTC to Eastern time.
                # EDT (spring-summer) is UTC-4. EST (fall-winter) is UTC-5
                # Exact dates for daylight savings is Mar 10 and Nov 3

                # First, check through all of the easy months.
                if months >= 4 and months <= 10: # EDT
                    hours -= 4
                elif months == 12 or months == 1 or months == 2: # EST
                    hours -= 5
                elif months == 3: # March.
                    if days <= 9: # EST
                        hours -= 5
                    else: # EDT
                        hours -=4
                elif months == 11: # November.
                    if days <= 3: # EDT
                        hours -= 4
                    else: # EST
                        hours -=5

                # Handle negative (or 0) days at the end, again
                if days <= 0:
                    months -= 1
                    if months in day31List:
                        days = 31 - abs(days)
                    elif months in day30List:
                        days = 30 - abs(days) 
                    else: # February: Not checking for leap years sorry
                        days = 28 - abs(days) 

                # Do some data conversion and then append the gps string. We use zfill to get the leading zeros back.
                entry = [str(hours).zfill(2), str(minutes).zfill(2), str(seconds).zfill(2), str(months).zfill(2), str(days).zfill(2)]
                timeStamps.append(entry)

        # Now we have an array of timestamps. Use GUI to ask the user for when the telescopes were started.
        layout = [[sg.Text("Look at these suggested start times from the GPS strings.")],
                  [sg.Text("Pick the one(s) that makes the most sense.")],
                  [sg.Text("Or just select nothing if they are all bad.")]]
        layout2 = []

        # Sort through the timestamps to get only unique ones? Potentially better to just pick
        # the most frequent timestamp and assume that's when the data collection started.
        uniqueTimes = [] # Holds the actual unique timestamps
        uniqueTimesCounts = [] # Holds the count of each unique timestamp
        seenValues = set()
        for item in timeStamps:
            t = tuple(item)
            if t not in seenValues:
                uniqueTimes.append(item)
                uniqueTimesCounts.append(int(1))
                seenValues.add(t)
            else:
                index = uniqueTimes.index(item)
                uniqueTimesCounts[index] += 1
                
        joined = []
        for x in range(len(uniqueTimes)):
            joined.append([uniqueTimes[x], uniqueTimesCounts[x]])

        joined = sorted(joined, key=lambda x: x[1], reverse=True)
        # The second "for" loop isn't really necessary, but it helps make the code more vertical
        # and not have too many characters on the same line
        formattedTimes = []
        amount = 0
        for values in joined:
            if amount < 12:
                theString = "Telescope recording began at (military time): "
                for x in range(len(values[0])):
                    if x < 2:
                        theString += values[0][x]
                        theString += ":"
                    elif x == 2:
                        theString += values[0][x]
                        theString += " and the date was: "
                    elif x == 3:
                        theString += values[0][x]
                        theString += "/"
                    else: # x >= 4
                        theString += values[0][x]

                theString += " (" + str(values[1]) + ")"       
                formattedTimes.append(theString) # append before concatenation
                layout2.append([sg.Checkbox(theString)])
                amount += 1
            else:
                print("ERROR: Too many different GPS strings for the program to display at once.\n")

        layout.append([sg.Frame("Possible telescope start times:", layout2, title_color="blue")])
        layout.append([sg.Submit()])
        window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
        event, values = window.Read()
        window.Close()

        for x in range(len(values)):
            if values[x] == True:
                selectGpsTimes.append(formattedTimes[x])
        
    return data, selectGpsTimes

# Does not work if a muon is not detected within the first second.
def find_start_index(lst):
    # Returns the index in the data when the time is zero (reset)
    # This will only keep going until the first gps reset in a file 
    index = 0
    for i in range(len(lst)):
        sec_timestamp = lst[i]
        if sec_timestamp[0] == 0:
            return i

    return 0

def correct_time(lst):
    global LM555FACTOR
    # If there is a jump in time, correct it
    # Look through the list and compare the times for three entries.
    # If the third timestamp is less than the second timestamp, then the 
    # third timestamp is probably in error - but look at a fourth timestamp first.
    # If the second timestamp is less than the first, AND it's been 
    # marked as a probable error, fix it by adopting the third timestamp's
    # seconds (sec) time
    # If the second timestamp is greater than the third, correct it by
    # adopting the first timestamp's seconds (sec) time
    # Example: 00001022, 00002032, 00001042 -> 00001022, 00001032, 00001042
    # Example: 00003022, 00003012, 00004022 -> 00003022, 00004012, 00004022
    # ADDED: If timeB < timeA, fix like in example 2
    # Removed while(true) loop because we should avoid correcting the data
    # too much.
    changes = 0
    for i in range(len(lst) - 3):
        timeA = lst[i][0] + lst[i][1]/LM555FACTOR
        timeB = lst[i+1][0] + lst[i+1][1]/LM555FACTOR
        timeC = lst[i+2][0] + lst[i+2][1]/LM555FACTOR
##        print(timeA, timeB, timeC)
        if ((timeB > timeA) and (timeC > timeB)):
            pass # everything is fine
        elif ((timeB > timeA) and (timeC < timeB)):
            # third timestamp is less than second timestamp
            # either because third timestamp is low or second timestamp
            # is high, so ....
            # check the fourth timestamp
            # This program is not intended to handle an errant 4th timestamp.
            timeD = lst[i+3][0] + lst[i+3][1]/255
            if (timeD < timeC):
                continue
            expected_timeC = 0.66*(timeD - timeA) + timeA
            diffC = expected_timeC - timeC
            expected_timeB = 0.33*(timeD - timeA) + timeA
            diffB = timeB - expected_timeB
            if (diffC > diffB) and (lst[i+2][0] != lst[i+3][0]):
                # third timestamp is too low compared to 4th timestamp
                lst[i+2][0] = lst[i+3][0]
                changes += 1
            elif (diffB > diffC) and (lst[i+1][0] != lst[i][0]):
                # second timestamp is too high
                lst[i+1][0] = lst[i][0]
                changes += 1
                
        # Added conditional below:
        elif ((timeB < timeA) and (timeC > timeB)):
            # timeB is too small. If possible, give timeB the time that timeA has.
            # If the subseconds are too low to make this reasonable, then use the
            # seconds from timeC instead.
            # If timeB seconds are lower and timeB subseconds are higher:
            if (lst[i][0] > lst[i+1][0]) and (lst[i][1] < lst[i+1][1]):
                lst[i+1][0] = lst[i][0]
                changes += 1
            else:
                lst[i+1][0] = lst[i+2][0]
                changes += 1

    print("The data was changed", changes, "times because of timeskip errors.")

    # Now that we are done correcting small skips, evaluate the data and tell the user
    # how many mistakes are left.
    # Again, we should avoid modifying the data too much.
    mistakes = 0
    for x in range(len(lst) - 1):
        if (lst[x][0] == lst[x + 1][0]):
            if (lst[x][1] > lst[x+1][1]):
                mistakes += 1
        else:
            timeA = lst[x][0] + lst[x][1]/LM555FACTOR
            timeB = lst[x + 1][0] + lst[x + 1][1]/LM555FACTOR
            if (timeB < timeA):
                mistakes += 1
    print("\nAfter correcting small time skips throughout the file,")
    print("there are still", mistakes, "mistakes left in the data. The program is not")
    print("intended to fix every mistake. If", mistakes, "is a large number,")
    print("perhaps the data is bad.\n")
    reportString = str(changes) + " changes were made to the data. There are still " + str(mistakes) + " mistakes left in the data."
    lst.insert(0, reportString)
    return lst 

# Function for 1)
# glob.glob DOES NOT sort the files properly right away
# THIS IS CASE INSENSITIVE: no files with f.txt allowed in the directory!
def combineHexFile(outputFileName, deleteHex):    
    print("Reading and sorting through appropriate files in", os.getcwd(), "...")
    # Get an array of file objects. They aren't just string names, or this could be less processing work.
    fileList = []
    for file in glob.glob(namePattern):
        fileList.append(file)

    # Error: no files found.
    if len(fileList) == 0:
        layout = [[sg.Text("Error: There's no files named 'F***.txt' here!. ", text_color="red")],
                  [sg.Text("Make sure you select the folder directory carefully.", text_color="red")],
                  [sg.Ok()]]

        window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
        event, values = window.Read()
        window.Close()
        raise FileNotFoundError # This would happen later automatically anyways.
        
    # Try organizing the list of files
    # By the numbers that follow the F letter
    fileNumbers = []
    for fileName in fileList:
        tempIndex = 1 # skip the first letter
        tempString = ""
        while ((fileName[tempIndex] != " ") and (fileName[tempIndex] != ".")):
            tempString += fileName[tempIndex]
            tempIndex += 1
        fileNumbers.append(int(tempString))

    # Now we have an array of the file numbers
    # Find their tags (0, 1, 2, etc) and append
    # to the new, final list based on tag magnitude
    finalList = []
    fileTag = 0
    while len(finalList) < len(fileNumbers):
        arrayIndex = 0
        # The following even collects multiple files with same tag!
        for entry in fileNumbers: 
            if fileTag == entry:
                finalList.append(fileList[arrayIndex])
            arrayIndex += 1
        fileTag += 1

    # Finally, we can start reading and writing files.
    filesRead = 0
    for file in finalList:
        # Create new list to avoid newline duplication!
        printList = []
        data1 = readfile(file)
        for data in data1:
                entry = data
                printList.append(entry)
        # print("file contains", len(times), "events")
        # print("Correcting time skips...")
        os.chdir(newDirectory)
        writefile(printList, outputFileName)
        os.chdir(baseDirectory)
        filesRead += 1

    print("Files read:", filesRead)
    print("Finished combining HEX files.")

    if deleteHex == True:
        for file in fileList:
            os.remove(os.getcwd() + "\\" + file)

# Function 2)
# This cannot use the same writefile function because we are writing a decimal file with 2 columns
def convHEXtoDEC(outputFileName, syncByGPSString):
    global LM555FACTOR
        
    print("Converting this HEX file to decimal:", "HEX" + outputFileName)
    if syncByGPSString:
        print("Attempting to sync the files based upon their GPS string.")
        data1, gpsdata = readHEXtoDEC_GPSSYNC("HEX" + outputFileName)
    else:
        data1, gpsdata = readHEXtoDEC("HEX" + outputFileName)
    print("Starting from the last GPS reset...")
    start_index = find_start_index(data1)
    if start_index is None:
        print("\n No GPS reset was found in this file!\n")
        telescope1 = data1
    else:
        telescope1 = data1[start_index:] 
    print("Correcting time skips...")
    telescope1 = correct_time(telescope1)
    
    # New: Record the amount of times that the data was collected and add it to the gps string list.
    gpsdata.insert(0, telescope1[0])
    telescope1 = telescope1[1:]
    
    printList = []
    lineNum = 0
    rollovers = 0
    try:
        for data in telescope1:
            decimal = data[1] / LM555FACTOR
            if decimal > 1 + (1/LM555FACTOR):
##                print("rollover:", data[1])
                rollovers += 1
            entry = data[0] + (data[1] / LM555FACTOR)
            printList.append(entry)
            lineNum += 1
    except ValueError:
        print("In one of the HEX files, there is an issue with a timestamp.")
        print("Sometimes, the HEX timestamp does not have all 8 digits. When")
        print("this happens, the program cannot interpret the data properly.")
        print("You will have to manually find the issue(s) in the file")
        print("and manually fix them.\n")
        print("THE ISSUE WAS ENCOUNTERED AT LINE NUMBER:", lineNum, "\nin the")
        print("combined HEX file. However, you will have to restart the program")
        print("after modifying the corresponding F0.txt files.")
        print("You can also figure out which F0.txt file has the problem by looking at the")
        print("nearby timestamps in the combined HEX file.\n")
        
    rolloverString = "There were " + str(rollovers) + " rollovers in the data where the subseconds were over "
    rolloverString += str(int(LM555FACTOR)) + " in HEX."
    print(rolloverString)
    print("Those rollovers were not deleted because they still represent real detections.")
    print("Note that we are not guaranteed to see every rollover because we rely on a detection coming in")
    print("while the counter is between 244 and 255.")
    fp = open(outputFileName, "a")
    print(rolloverString, file=fp)
    for string in gpsdata:
##        print(string)
        print(string, file=fp)
    
    for entry in printList: 
        print(entry, file=fp)

    fp.close()
    print("Finished converting to decimal. There were", len(printList), "events.")
    print("")

# master is a list like:
# [seconds, telescopeNum] 
# Scan through every time in the master array. In this master array, we are looking
# for coincident events that can only be separated by a small time duration. We call
# this time duration the scan_window
def scan_times(master, window):
    global LM555FACTOR # to be able to access the global variable
    # Recursive function for scanning through successively larger windows.
##    for i in master:
##        print(i)
    def scan_within_window(master, window, maxWindow, coincidences):
        global LM555FACTOR
        if window > maxWindow or len(master) <= 1:
            return coincidences
        scan_window = (window / LM555FACTOR) + (0.0001) # Add small amount to scan_window to prevent floating point errors.
        # Note that the timestamps and their differences are quantized anyways. Adding this small amount is
        # physically inconsequential, but it makes this scanning algorithm run as intended even if there
        # are small errors in the timestamp decimals.

        # First check to see if the current coincidences can be enlarged.
        # Then, collect any leftover coincidences that are not extensions of smaller coincidences.      
        coincident_telescopes = []
        updated_coincidences = []
        coincidence_pop_indices = []
        time_diff = 0
        # TODO: Doesn't obey window restriction
        if len(coincidences) > 0:
            for i in range(len(coincidences)):
                start_time = coincidences[i][1]
                end_time = coincidences[i][2] + start_time
                x = 0
                pop_indices = []
                while x < len(master):
                    # Break condition to save iteration processing time.
                    if master[x][0] - start_time > scan_window: # apply upper limit on timestamp
##                        updated_coincidences.append(coincidences[i])
                        # Do not need to pop the master element here.
                        break
                    # check to see if the detection is within scan_window from start_time and end time
                    # no need for abs()
                    elif master[x][1] not in coincidences[i][3]: # Telescope is new to the coincidence
                        if end_time - master[x][0] <= scan_window: # apply lower limit on timestamp
                            # This is coincident. Modify the coincidence.
                            # Second part of conditional can be commented because it is part of the break condition.
                            if master[x][0] < start_time:
                                start_time = master[x][0]
                                time_diff = end_time - start_time
                            else:
                                time_diff = master[x][0] - start_time
                            
##                            coincident_telescopes = coincidences[i][3] # Bad. This is only reference assignment
                            coincident_telescopes = coincidences[i][3][:] # Fastest copy method using full-list slice
                            
                            coincident_telescopes.extend([master[x][1]]) # the extra brackets are needed
                            entry = [len(coincident_telescopes), start_time, time_diff, coincident_telescopes]
                            updated_coincidences.append(entry)
                            pop_indices.append(x)
                            if i not in coincidence_pop_indices:
                                coincidence_pop_indices.append(i)
                            # break # Do not break here because you could catch more than 1 timestamp
                    x += 1
                x = len(pop_indices) - 1
                while x >= 0:
                    master.pop(pop_indices[x])
                    x -= 1

        x = len(coincidence_pop_indices) - 1
        while x >= 0:
            coincidences.pop(coincidence_pop_indices[x])
            x -= 1
        
        for i in range(len(coincidences)):
            updated_coincidences.append(coincidences[i])

        updated_coincidences = sorted(updated_coincidences, key=lambda x: x[1])
        for i in updated_coincidences:
            print(i)
        input("Above is before the function")

        # Check to see if adjacent coincidences can be combined. It's good to do this while the window is
        # as small as possible.
        # We use this slicing notation: [:] to quickly copy a list by value and not by reference.
        start_time = 0 # Holds the earliest time from a potential coincidence
        end_time = 0 # Holds the latest time from a potential coincidence from previous iteration
        current_time = 0 # Holds the time from the current iteration
        time_diff = 0 # Holds the time difference from previous iteration
        pop_indices = []
        temp_indices = [0]
        if len(updated_coincidences) > 1:
            previous_telescopes = updated_coincidences[0][3][:]
        x = 0
        while x < len(updated_coincidences) - 1:
            current_time = updated_coincidences[x+1][1]
            current_telescopes = updated_coincidences[x+1][3]
            conditionalList = [current_telescopes[z] in previous_telescopes for z in range(len(current_telescopes))]
##            if True not in conditionalList:
##                print(current_time, start_time, end_time)
##                print(updated_coincidences[x+1], updated_coincidences[x])
            if current_time - start_time <= scan_window and True not in conditionalList:
                previous_telescopes.extend(current_telescopes)
##                start_time = min(start_time, current_time) # useless line
                end_time = current_time + updated_coincidences[x+1][2]
                temp_indices.append(x+1)
            else:
                if len(temp_indices) > 1: 
                    time_diff = end_time - start_time
                    if time_diff <= scan_window: # This coincidence is good to save!
                        entry = [0, start_time]
                        for k in range(len(temp_indices)):
                            entry[0] += updated_coincidences[x-k][0]
                            entry[1] = min(entry[1], updated_coincidences[x-k][1])
                        entry.append(end_time - start_time)
                        entry.append(previous_telescopes)
                        updated_coincidences.append(entry)

                        # To carefully pop from the coincidences list, iterate over the temp_indices list backwards.
                        c = len(temp_indices) - 1
                        while c >= 0:
                            if updated_coincidences[c][1] >= 15.80 and updated_coincidences[c][1] <= 15.81:
                                print(entry, x)
                                input("Attempt to delete the timestamp occurred here.")
                            updated_coincidences.pop(temp_indices[c])
                            x -= 1 # Since we removed a previous entry, have to reduce the index.
                            c -= 1
                previous_telescopes = current_telescopes[:]
                start_time = current_time
                temp_indices = [x+1] # Different from the append argument because this is before increment of x.
            x += 1

        # Catch the last coincidence combination. This is possible if the very last element
        # was not in conditionalList.
        if len(temp_indices) > 1:
            if time_diff <= scan_window:
                # This coincidence is good to save!
                entry = [0, start_time]
                for k in range(len(temp_indices)):
                    entry[0] += updated_coincidences[x-k][0]
                    entry[1] = min(entry[1], updated_coincidences[x-k][1])
                entry.append(end_time - start_time)
                entry.append(previous_telescopes)
##                        print(entry)
                updated_coincidences.append(entry)
                
                for val in temp_indices:
                    updated_coincidences.pop(val)

        updated_coincidences = sorted(updated_coincidences, key=lambda x: x[1])
        for i in updated_coincidences:
            print(i)
        input("Above is after the function")
        
        coincident_telescopes = []
        start_time = 0 # Holds the earliest time from a potential coincidence
        end_time = 0 # Holds the latest time from a potential coincidence from previous iteration
        current_time = 0 # Holds the time from the current iteration
        time_diff = 0 # Holds the time difference from previous iteration
        pop_indices = []
        temp_indices = []
  
        # This is a not-generous scanning algorithm, and this is being paired with the coincidence-expansion
        # algorithm. It's possible that the generous algorithm is better here?
        for i in range(len(master)):
##            print(master[i], coincident_telescopes)
            current_time = master[i][0] # in seconds
            current_telescope = master[i][1] # A number that is from 0 to the number of telescopes - 1
##            print(coincident_telescopes, current_time, start_time, current_time-start_time, scan_window)
##            print(current_telescope not in coincident_telescopes)
##            print(current_time - start_time <= scan_window, current_time - start_time, current_time - start_time - scan_window)
            if current_telescope not in coincident_telescopes and current_time - start_time <= scan_window:
                coincident_telescopes.append(current_telescope)
##                start_time = min(start_time, current_time) # useless line
                end_time = current_time
                temp_indices.append(i)
            else:
                if len(coincident_telescopes) > 1:
                    time_diff = end_time - start_time
                    if time_diff <= scan_window:
                        # This coincidence is good to save!
                        entry = [len(coincident_telescopes), start_time, time_diff, coincident_telescopes]
##                        print(entry)
                        updated_coincidences.append(entry)
                        for val in temp_indices:
                            pop_indices.append(val)
                coincident_telescopes = [current_telescope]
                start_time = current_time
                temp_indices = [i]
            
        # Catch the last coincidence too.
        if len(coincident_telescopes) > 1:
            updated_coincidences.append([len(coincident_telescopes), start_time, time_diff, coincident_telescopes])
            for val in temp_indices:
                pop_indices.append(val)

        # Make the master list shorter because we already used some of the timestamps.
        x = len(pop_indices) - 1
        while x >= 0:
            master.pop(pop_indices[x])
            x -= 1

        return scan_within_window(master, window + 1, maxWindow, updated_coincidences)
        # END RECURSIVE FUNCTION
        
    coincidence_list = []
    coincidence_list = scan_within_window(master, 0, window, coincidence_list)
    # Now make sure to sort the result by the timestamp
    coincidence_list = sorted(coincidence_list, key=lambda x: x[1])
    # TODO: this recursive function did not check to see if coincidences can be combined.    
    return coincidence_list


def createLightCurve(outputFileName, tempDuration, genIntervals, maxIntervalDuration):
    global LM555FACTOR
    global ArduinoDeadTime
    global BS2DeadTime
    # Assumes that the file is a decimal file, without any separation between seconds and subseconds
    # Therefore only capatible with data from 2019 summer and after, unless that data is converted back
    # to hex and then back to decimal.
    print("Generating light curve in excel...")
    
    # Work with the user input, and don't let the program crash if the input is bad.
    if tempDuration is not "":
        try:
            binDuration = float(tempDuration)
        except ValueError:
            print("The bin duration didn't make sense, so we will just use the default bin duration of")
            print(binDuration, "seconds.")
            pass
        
    data = readfile(outputFileName) # Get all of the timestamps from a file, but they are in string format.
##    for i in data:
##        print(i)
    timeStamps = []
    index = 0
    # Find all of the timestamps by using ValueError exceptions
    for x in range(len(data)):
        try:
            test = float(data[x])
        except ValueError:
            # This must have been a timestamp.
            timeStamps.append(data[x])
            index = x + 1
            # don't break yet
        else:
            break

    data = data[index:]

    # Now there are 2 arrays we have: data[] and timeStamps[]
    # data is now just the decimal times. timeStamps is just the selected gps strings.
    data = [float(i) for i in data] # So convert all of these strings to float values.

    timeDuration = max(data) # Just use the largest timestamp for the binrange
    maxBinRow = int(timeDuration / binDuration) # The last bin is cut off intentionally.
    binValues = [(i+1)*binDuration for i in range(maxBinRow)]

    # Now do the count sorting based upon how the times compare to the bin ranges
    counts = [0 for i in range(maxBinRow)] # Start this as an array of 0's whose values will be changed anyways

    for numbers in data:
        binNum = int(numbers / binDuration)
        if binNum < maxBinRow:
            counts[binNum] += 1
    
    workbook = Workbook() # Create workbook object reference
    worksheet = workbook.active # Create current excel sheet object reference (just to change the title)
    worksheet.title = "Light Curve"
    worksheet.append(["Time (s)", "Bin (s)", "Bin Counts", "Adjusted Sq. Residuals"]) # Column titles
    worksheet.column_dimensions['C'].width = 9.71 # Make C column wider for "Bin Counts"

    # Add all data to the worksheet
    for i in data:
        worksheet.append([float(i)])

    # If the time doesn't go on long enough to make a bin, then return now to save time.
    # The user will only have a column of timestamps, and nothing else.
    if maxBinRow <= 0:
        workbook.save(outputFileName[:-4] + ".xlsx")
        print("There was not enough data to create a full-sized bin. Try using a smaller bin size.")
        print("The program will now exit.\n")
        return
    
    # Add the bin values next to it
    for x in range(maxBinRow):        
        worksheet.cell(row=x+2, column=2).value = binValues[x]
        worksheet.cell(row=x+2, column=3).value = counts[x]
        worksheet.cell(row=x+2, column=4).value = "=IF(C" + str(x + 2) + ">G$9/4, (C" + str(x+2) + " - G$10)^2, 0)"

    # Compute a few statistics for the user. Make them in orange color to caution the user
    # Against blindly trusting these numbers - if there are mistakes in the data or in the
    # Excel plot, then these statistics may be affected. For these reasons, certain statistics
    # Like std. deviation will not be plotted.
    # Before we do any of this, also make a warning for the user
    if len(timeStamps) > 0:
        for x in range(len(timeStamps)):
            worksheet.cell(row=x+1, column=9).value = timeStamps[x]
    worksheet.merge_cells("F2:G2")
    if min(counts) == 0:
        worksheet.cell(row=2, column=6).value = "Some bins have low or 0 counts!!!"
    else:
        worksheet.cell(row=2, column=6).value = "Sample statistics"

    worksheet.column_dimensions['D'].width = 20 # Wider column for the title 
    worksheet.column_dimensions['F'].width = 23.86 # Make the column wider for text
    worksheet.column_dimensions['G'].width = 12 # Make the column wider for numbers
    worksheet.cell(row=3, column=6).value = "Observation time:"
    worksheet.cell(row=4, column=6).value = "Total counts:"
    worksheet.cell(row=5, column=6).value = "Count rate:"
    worksheet.cell(row=6, column=6).value = "Dead time:"
    worksheet.cell(row=7, column=6).value = "True count rate:"
    worksheet.cell(row=9, column=6).value = "Bin Counts mean:"
    worksheet.cell(row=10, column=6).value = "Adjusted Bin Counts mean:"
    worksheet.cell(row=11, column=6).value = "Total bins:"
    worksheet.cell(row=12, column=6).value = "Bin Std Deviation:"
    
    worksheet.cell(row=5, column=7).value = "=G4 / G3"

    # Scoping rules: deadTime is a local variable available to this whole createLightCurve() function
    # and it will be used later for intervals.
    if LM555FACTOR == 255:
        deadTime = BS2DeadTime
    else:
        deadTime = ArduinoDeadTime
        
    worksheet.cell(row=6, column=7).value = deadTime
    worksheet.cell(row=7, column=7).value = "=G5 / (1 - (G5 * G6))"
    worksheet.cell(row=9, column=7).value = "=AVERAGE(C2:C" + str(maxBinRow + 1) + ")"

    # ONLY COUNTS CELLS IF THEY ARE GREATER THAN 1/4 OF THE MEAN
    worksheet.cell(row=3, column=7).value = "=" + str(binDuration) + " * COUNTIF(C2:C" + str(maxBinRow + 1) + ''', ">"&G9/4)'''
    worksheet.cell(row=4, column=7).value = "=SUMIF(C2:C" + str(maxBinRow + 1) + ''', ">"&G9/4)'''
    worksheet.cell(row=10, column=7).value = "=G4 / G3 * " + str(binDuration)
    worksheet.cell(row=11, column=7).value = "=COUNTIF(C2:C" + str(maxBinRow + 1) + ''', ">"&G9/4)'''
    worksheet.cell(row=12, column=7).value = "=SQRT(SUMIF(D2:D" + str(maxBinRow + 1) + ''', ">"&G9/4)''' + " / (G11 - 1))"
    

    # Create and format the Scatter Plot. Then save the worksheet.
    lightCurve = ScatterChart()
    lightCurve.title = outputFileName[:-4] # Hopefully the user used the telescope name as the filename...
    lightCurve.style = 5 # Just changes color/thickness of the line.
    lightCurve.x_axis.title = "Time (s)"
    lightCurve.y_axis.title = "Counts"
    lightCurve.y_axis.scaling.min = 0 # Make sure the y-axis starts from 0.
    lightCurve.x_axis.scaling.min = 0 # Same for x-axis.
    
    xvalues = Reference(worksheet, min_col=3, min_row=2, max_row=maxBinRow + 1)
    yvalues = Reference(worksheet, min_col=2, min_row=2, max_row=maxBinRow + 1)
    
    series = Series(xvalues, yvalues, title_from_data=False)
    lightCurve.series.append(series)                                                                            
    worksheet.add_chart(lightCurve, "J3")

    # ===== Sheet 2 =====
    if genIntervals:
        print("Generating interval graphs...")
        worksheet = workbook.create_sheet("Intervals")

        # Create column titles
        worksheet.append(["DeadTime-Adjusted Time Separations (s)", "Bin (s)", "Counts", "Ln(Counts)", "Residuals"])
        worksheet.column_dimensions['A'].width = 36.5
        worksheet.column_dimensions['D'].width = 9.8
        
        # Just compute the intervals directly. Better than copying the long timestamps list again.
        intervalList = []
        for x in range(len(data) - 1):
            entry = data[x+1] - data[x] - deadTime
            worksheet.append([entry]) # Yes, it has to be in brackets for this append method.
            intervalList.append(entry) 
        
        intervalBinSize = 1 / 244.1 # about 4ms.
        maxHistogramRow = int(maxIntervalDuration / intervalBinSize)
        binValues = [(i+1)*intervalBinSize for i in range(maxHistogramRow)]

        # Do the count sorting based upon the intervals compared to bin ranges.
        intervalHistogram = [0 for i in range(maxHistogramRow)] # An array of 0's whose values will be incremented.

        # Note that NEGATIVE INTERVALS HAVE NOT BEEN COUNTED, DIFFERENT FROM THE EXCEL HISTOGRAM
        # Therefore if you compare this output to the output of excel's histogram function, then
        # the first bin produced by excel will have higher counts in it if there are any negative intervals
        # in your dataset (mistakes).
        # This program will not count those negative intervals in the first bin.
        for numbers in intervalList:
            # binNum = int(numbers / intervalBinSize) # This is the most intuitive way to do it
            if numbers >= 0:
                binNum = math.ceil(numbers / intervalBinSize) - 1 # Do it this way
                if binNum < maxHistogramRow:
                    intervalHistogram[binNum] += 1
##            else:
                # Do something with the negative intervals here if needed

        if len(intervalHistogram) > 0:
            for x in range(maxHistogramRow):
                worksheet.cell(row=x+2, column=2).value = binValues[x]
                worksheet.cell(row=x+2, column=3).value = intervalHistogram[x]

        # Make the scatter plot too.
        intervalCurve = ScatterChart()
        intervalCurve.title = "Detection Time Separations"
        intervalCurve.style = 12
        intervalCurve.x_axis.title = "Time Separation (s)"
        intervalCurve.y_axis.title = "Counts"
        intervalCurve.y_axis.scaling.min = 0 # Make sure the y-axis starts from 0.
        intervalCurve.x_axis.scaling.min = 0 # Same for x-axis.

        xvalues = Reference(worksheet, min_col=2, min_row=2, max_row=maxHistogramRow + 1)
        yvalues = Reference(worksheet, min_col=3, min_row=2, max_row=maxHistogramRow + 1)
        series = Series(yvalues, xvalues, title_from_data=False)
        intervalCurve.series.append(series)                                              
        worksheet.add_chart(intervalCurve, "J3")
        
        # Now do the log-scaled counts
##        logValues = []
        for x in range(len(intervalHistogram)):
            if intervalHistogram[x] > 0:
                value = math.log(intervalHistogram[x])
                worksheet.cell(row=x+2, column=4).value = value
##                logValues.append(value)
##            else:
##                worksheet.cell(row=x+2, column=4).value = 0 # This actually messes with models for the data.
##                logValues.append(0)

        # Make a separate log-scaled graph
        logIntervalCurve = ScatterChart()
        logIntervalCurve.title = "Log-Scaled Detection Time Separations"
        logIntervalCurve.style = 12
        logIntervalCurve.x_axis.title = "Time Separation (s)"
        logIntervalCurve.y_axis.title = "LN(Counts)"
        logIntervalCurve.y_axis.scaling.min = 0 # Make sure the y-axis starts from 0.
        logIntervalCurve.x_axis.scaling.min = 0 # Same for x-axis.        
        
        xvalues = Reference(worksheet, min_col=2, min_row=2, max_row=maxHistogramRow + 1)
        yvalues = Reference(worksheet, min_col=4, min_row=2, max_row=maxHistogramRow + 1)
        series = Series(yvalues, xvalues, title_from_data=False)
        logIntervalCurve.series.append(series)                                                                            
        worksheet.add_chart(logIntervalCurve, "J22")

        # The next step would be to begin a linear regression for the log-scaled data...
        # Numpy has a built-in function for this.
##        logValues = np.array(logValues)
##        xValues = np.array(binValues)
##        b, m = polyfit(xValues, logValues, 1) # "1" argument makes this first order regression

##        # TODO: Now store these values somewhere, and compute columns for the linear model
##        worksheet.cell(row=15, column=6).value = "Linear Model"
##        worksheet.cell(row=16, column=6).value = "m"
##        worksheet.cell(row=17, column=6).value = "b"
##        worksheet.cell(row=18, column=6).value = "SSR"
##        worksheet.cell(row=19, column=6).value = "Total Squares"
##        worksheet.cell(row=20, column=6).value = "R²"
##        
##        
##        worksheet.cell(row=16, column=7).value = m
##        worksheet.cell(row=17, column=7).value = b
##        worksheet.cell(row=18, column=7).value = "=SUM(D2:D" + str(maxHistogramRow + 1) + ")"
##        worksheet.cell(row=19, column=7).value = "=SUM(
##        worksheet.cell(row=20, column=7).value = "SSR"
##        worksheet.cell(row=21, column=7).value = "Total Squares"
##        worksheet.cell(row=20, column=7).value = "R²"
        
    workbook.save(outputFileName[:-4] + ".xlsx")
    print("Finished generating light curve,")
    print("and saved it to:", outputFileName[:-4], ".xlsx")
    print("There are", maxBinRow, "points on the graph.")

# Requires the user to be in the working directory of the data, so that OS can check to see
# if that folder already exists.
def getOutputFolder():
    # Use GUI to get the output file name from the user.
    # Current date is obtained from computer.
    today = str(date.today())
    today = str(int(today[5:7])) + today[-3:]

    layout = [[sg.Text("This program will send the output files to a new folder.")],
              [sg.Text("Select a name for the output to be sent.")],
              [sg.Text("For example, write: "), sg.Text(today + " 201 & 202", text_color='red')],
              [sg.InputText(), sg.Text(".txt")],
              [sg.Submit()]]
    window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
    event, values = window.Read()
    window.Close()
    
    # This time the return value is just a 1-element array of the single input string,
    # but it still needs manual indexing for access.
    outputFileName = values[0] + ".txt"
    while (os.path.exists(os.getcwd() + outputFileName[:-4])):
        layout = [[sg.Text("THE FOLDER: ", text_color="red"), sg.Input(default_text=outputFileName[:-4], disabled=True)],
                  [sg.Text("ALREADY EXISTS IN: ", text_color="red"), sg.Input(default_text=os.getcwd(), disabled=True)],
                  [sg.Text("PLEASE MOVE OR DELETE IT!", text_color="red")],
                  [sg.Ok()]]

        window = sg.Window("Analyze Cosmic Ray Telescope Data", layout)
        window.Read()
        window.Close()
        
        layout = [[sg.Text("This program will send the output files to a new folder.")],
                  [sg.Text("Select a location for the output to be sent.")],
                  [sg.Text("For example, write: "), sg.Text(today + " 201 & 202", text_color='red')],
                  [sg.InputText(), sg.Text(".txt")],
                  [sg.Submit()]]      
        window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
        event, values = window.Read()
        window.Close()

        outputFileName = values[0] + ".txt"
    return outputFileName
#==================================================#
#                   MAIN PROGRAM                   #
#==================================================#

# Infinite loop allowing for faster consecutive use.
# The first two cases were fit into functions, while 3 and 4 are written explicitly below.
##while 1:
layout = [[sg.Text("Select an option:")],
        [sg.Radio("Combine HEX files", "RADIO1", default=True), sg.Radio("Coincidence Scanning", "RADIO1"),
         sg.Radio("Convert a decimal file back into a HEX file", "RADIO1")],
        [sg.Submit()]] 

window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
event, values = window.Read()
window.Close()

# For radio GUI objects, the 'values' return array is special (must use manual indexing to get return values)
# Find out which of the options was selected, and store it into a variable that can be interpreted more easily
choice = 0
# Assign a value of 1, 2, or 3 to "choice"
for x in range(len(values)):
    if values[x] == True:
        choice = x + 1

# Code for the three different functions of this program below.
if (choice == 1):
    tempString = sg.PopupGetFile("Open one of the hex files for the program to start.", default_path="C:\\") #for anticoincidence

    # Now delete all of the hex filename text, up to the last '/'
    # Note: the os. library expects backward slashes '\' and not forward slashes '/'
    # So we can't do equality tests with os.getcwd() and the folder names we obtain
    # Fortunately, os.chdir() doesn't care about the directionality of the slashes.
    folder = ""
    # "partition("/") does not work for the string on the line below, so "split" is used instead
    # "C:/Users/alexl/Downloads/Unsynced Zenith Data/6-21 Zenith start 10am/6-21 203 & 204 start 10am/F0.txt"
    tempList = tempString.split("/") 
    i = 0
    while i < len(tempList) - 1:
        folder += tempList[i]
        folder += "/"
        i += 1
    os.chdir(folder)
    print("Ok, we will be working with HEX files from this folder:", os.getcwd())
    print("Files that do not follow the pattern 'F*.txt' will be ignored.")
    print("'*' is a wildcard symbol.")

    # Careful about changing these lines, these are specifically for windows and mac compatibility
    baseDirectory = os.getcwd() # Which is equal to folder, but with all of the "/" reversed
    outputFileName = getOutputFolder() # Use GUI to prompt user for new folder name
    newDirectory = baseDirectory + "/" + outputFileName[:-4]
    os.mkdir(newDirectory)

    layout = [[sg.Text("Is this data from the Arduino?", size=(25,1)), sg.Radio("Yes", "RADIO1", default=True), sg.Radio("No", "RADIO1")],
              [sg.Text("Delete the small, older HEX files?", size=(25,1)), sg.Radio("Yes", "RADIO2", default=True), sg.Radio("No", "RADIO2")],
              [sg.Text("Generate Light Curve?", size=(25,1)), sg.Radio("Yes", "RADIO3", change_submits=True, default=True), sg.Radio("No", "RADIO3", change_submits=True), sg.Text("Bin Size (s):"), sg.Input("337.5", change_submits=True)],
              [sg.Text("Generate Interval Statistics?", size=(25,1)), sg.Radio("Yes", "RADIO4", change_submits=True, default=True), sg.Radio("No", "RADIO4", change_submits=True), sg.Text("Max Interval Duration (s):"), sg.Input("5", change_submits=True)],
              [sg.Text("New: Use GPS times to detect rollovers?", size=(25,1)), sg.Radio("Yes", "RADIO5", change_submits=True), sg.Radio("No", "RADIO5", change_submits=True, default=True)],
              [sg.Submit()]]

    window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
    event, values = window.Read()

    oldString = "" # Just needs to be defined for this while loop
    oldString1 = "" # Needs to be defined
    # We use this while-loop to prevent the user from entering anything besides a float number in the box.
    while event is not None and event and event is not "Submit":
        genLightCurve = values[4]
        if genLightCurve == False:
##                window.FindElement(7).Update(visible=False)
            window.FindElement(6).Update(disabled=True) # Disable input box
        else:
##                window.FindElement(5).Update(visible=True)
            window.FindElement(6).Update(disabled=False) # Enable input box
        
        genIntervals = values[7]
        if genIntervals == False:
            window.FindElement(9).Update(disabled=True) # Disable input box
        else:
            window.FindElement(9).Update(disabled=False) # Enable input box

        # Now take advantage of the old values stored in newString and oldString1 to delete any
        # secondary period. newString is for values[6], oldString1 is for values[9].
        # Obviously this won't work on the first iteration of the while loop, so we do None/Null checking.
        # Unfortunately the text cursor will move right if a second period is typed.
        if "." in oldString:
            periodCount = 0
            for x in range(len(values[6])):
                if values[6][x] == ".":
                    periodCount += 1
                if periodCount > 1:
                    values[6] = oldString
                    break

        if "." in oldString1:
            periodCount = 0
            for x in range(len(values[9])):
                if values[9][x] == ".":
                    periodCount += 1
                if periodCount > 1:
                    values[9] = oldString1
                    break
    
        # Restrict text box input to just numbers.
        # Also have to delete multiple periods.
        # Text box values are stored in values[6] and values[9]
        if len(values[6]) > 0:
            oldString = ""
            for x in range(len(values[6])):
                # This makes it easier to write the conditional:
                chrVal = ord(values[6][x]) # Conversion from chr to int. "0" = 48 and "9" = 57.
                if values[6][x] == "." or (chrVal >= 48 and chrVal <= 57):
                    oldString += values[6][x]
            window.FindElement(6).Update(value=oldString)

        # Same loop as before but with values[9]  
        if len(values[9]) > 0:
            oldString1 = ""
            for x in range(len(values[9])):
                # This makes it easier to write the conditional:
                chrVal = ord(values[9][x]) # Conversion from chr to int. "0" = 48 and "9" = 57.
                if values[9][x] == "." or (chrVal >= 48 and chrVal <= 57):
                    oldString1 += values[9][x]
            window.FindElement(9).Update(value=oldString1)
        event, values = window.Read()

    window.Close()

    # Unfortunately, the values return array is only 1D.
    # So we have to be smart about how to check each element
    # Here is the structure of the return array:
    # values[0]  Data is from Arduino
    # values[1]  Data is not from Arduino
    # values[2]  Delete old hex files
    # values[3]  Keep old hex files
    # values[4]  Generate light curve
    # values[5]  Do not generate light curve
    # values[6]  Bin Duration value for light curve
    # values[7]  Generate interval statistics
    # values[8]  Do not generate interval statistics
    # values[9]  Maximum time separation value for interval statistics. It is imperative to
    #            always use the bin duration of 1/244.1 seconds as the intervals are quantized this way,
    #            so the bin size is not an option for the user.
    # values[10] Use GPS strings to detect rollovers
    # values[11] Do not use GPS strings to detect rollovers
    # ...
    # Note that values[0] and values[1] are not independent. The same is true for values[2]
    # and values[3], 4 and 5, 7 and 8, 10 and 11.

    # Booleans taken from GUI interface. However Python scoping rules means we didn't need to do this.
    if values[0] == True:
        LM555FACTOR = 244.1 # Frequency of new crystal oscillator
    else:
        LM555FACTOR = 255   # Old BS2 LM555 frequency        
    deleteHex = values[2]
    genLightCurve = values[4] # Added for completeness
    genIntervals = values[7] # Also for completeness
    syncByGPSString = values[10]

    # Below are the numerical inputs from the GUI interface. We have to use try statements because
    # PySimpleGUI doesn't allow restricting inputs to numbers.
    tempDuration = float(values[6]) # Number written to first text box.
    maxIntervalDuration = float(values[9]) # Number written to second text box.
    
    # Directory switching happens between every file in combineHexFile, inevitably.
    # There is no such thing as "moving" the files without copying and then deleting.
    combineHexFile("HEX" + outputFileName, deleteHex)

    os.chdir(newDirectory)
    convHEXtoDEC(outputFileName, syncByGPSString)
    if genLightCurve:
        createLightCurve(outputFileName, tempDuration, genIntervals, maxIntervalDuration)

elif (choice == 2):
    layout = [[sg.Text("For this function to work, you need to copy the decimal files into")],
             [sg.Text("the same folder. Then, the program will just use all of the files")],
             [sg.Text("in that folder for the coincidence detection. There should be no other")],
             [sg.Text("files inside of that folder, or else this function will fail.")],
             [sg.Text("Open one of the files inside of that folder:", text_color="red")],
             [sg.Input(), sg.FileBrowse()],
             [sg.Submit()]]

    window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
    event, values = window.Read()
    window.Close()
    tempString = values[0]
    # Now delete all of the hex filename text, up to the last '/'
    folder = ""
    # "partition("/") does not work for the string on the line below, so "split" is used instead
    # "C:/Users/alexl/Downloads/Unsynced Zenith Data/6-21 Zenith start 10am/6-21 203 & 204 start 10am/F0.txt"
    tempList = tempString.split("/") 
    i = 0
    while i < len(tempList) - 1:
        folder += tempList[i]
        folder += "/"
        i += 1
    os.chdir(folder)
    print("Ok, we will be working with HEX files from this folder:", os.getcwd())

    # Careful about changing these lines, they are for windows-mac compatibility
    baseDirectory = os.getcwd() # Which is equal to folder, but with all of the "/" reversed
    outputFileName = getOutputFolder() # Use GUI to prompt user for new folder name
    newDirectory = baseDirectory + "/" + outputFileName[:-4]
    os.mkdir(newDirectory)
    
    # Use GUI to ask the user for each telescope's "offset," in case the telescopes weren't manually synced.
    layout = [[sg.Text("Input the relative offset of each telescope in seconds")],
              [sg.Text("(this number can be a decimal). Note that these numbers are relative,")],
              [sg.Text("so if you add time to every telescope then no offset has actually been")],
              [sg.Text("instituted for one telescope. I recommend choosing only positive numbers")],
              [sg.Text("so that T=0 starts from when the last telescope began recording.")]]
    num = 0
    layout2 = []
    for file in glob.glob("*.txt"):
        layout2.append([sg.Text(file), sg.Input("0", change_submits=True, size=(15, 1))])

    layout.append([sg.Frame("Telescope List", layout2, title_color="blue")])
    layout.append([sg.Submit()])

    window = sg.Window("Analyze Cosmic Ray Telescope Data", layout)
    event, values = window.Read()
    oldStrings = ["" for x in range(len(values))]

    while event is not None and event is not "Submit":
        # Only allow 1 period.
        for y in range(len(oldStrings)):
            if "." in oldStrings[y]:
                periodCount = 0
                for x in range(len(values[y])):
                    if values[y][x] == ".":
                        periodCount += 1
                    if periodCount > 1:
                        values[y] = oldStrings[y]
                        break

        # Only allow one minus sign, but it also has to be the first character.
        for y in range(len(oldStrings)):
            if "-" in values[y]:
                if values[y][0] != "-":
                    values[y] = oldStrings[y]
                else:
                    # Also make sure there isn't a second minus sign.
                    minusCount = 0
                    for x in range(len(values[y])):
                        if values[y][x] == "-":
                            minusCount += 1
                        if minusCount > 1:
                            values[y] = oldStrings[y]
                
                
            # Restrict text box input to just numbers, periods, and minus signs.
            # Also have to delete multiple periods.
            # Text box values are stored in values[6] and values[9]
            if len(values[y]) > 0:
                oldStrings[y] = ""
                for x in range(len(values[y])):
                    # This makes it easier to write the conditional:
                    chrVal = ord(values[y][x]) # Conversion from chr to int. "0" = 48 and "9" = 57.
                    if values[y][x] == "." or (chrVal >= 48 and chrVal <= 57) or chrVal == 45:
                        oldStrings[y] += values[y][x]
                window.FindElement(y).Update(value=oldStrings[y])
            
        event, values = window.Read()
    window.Close()

    # Extract the offsets.
##        offsets = []
##        for x in range(len(values)):
##            offsets.append(float(values[x]))

    # Also ask if the data was from the Arduino.
    layout = [[sg.Text("Is this data from the Arduino?")],
              [sg.Radio("Yes", "RADIO1", default=True), sg.Radio("No", "RADIO1")],
              [sg.Text("Scan window:"), sg.Input("1", change_submits=True, size=(15, 1))],
              [sg.Submit()]]

    window = sg.Window("Analyze Cosmic Ray Telescope Data", layout)
    event, values = window.Read()
    oldString = values[2]
    while event is not None and event is not "Submit":
        # Restrict text box input to just integers
        # Text box value is stored in values[2]
        if len(values[2]) > 0:
            oldString = ""
            for x in range(len(values[2])):
                # This makes it easier to write the conditional:
                chrVal = ord(values[2][x]) # Conversion from chr to int. "0" = 48 and "9" = 57.
                if chrVal >= 48 and chrVal <= 57:
                    oldString += values[2][x]
            window.FindElement(2).Update(value=oldString)
        event, values = window.Read()
    window.Close()
    
    if values[0] == True:
        LM555FACTOR = 244.1 # Frequency of new crystal oscillator
    else:
        LM555FACTOR = 255   # Old BS2 LM555 frequency

    selected_window = int(values[2])
    print("using LM555FACTOR of:", LM555FACTOR)
    for k in range(0, 1):
        offsets = [0, 0, 0, 0] # DELETE THIS
##            print("Obtained offsets.") 
        
        # Now make the list of all timestamps that will be sorted after. This assumes
        # the user put the telescope files all in the same folder.
        master = []
        telescopeNum = 0
        for file in glob.glob("*.txt"):
            data = readfile(file)
            for entry in data:
                try:
                    master.append([float(entry) + offsets[telescopeNum], telescopeNum])
                except ValueError:
                    # Assume that it was just a GPS string or some other string
                    continue
            telescopeNum += 1
            
        master1 = sorted(master, key=lambda x: (x[0], x[1])) # Now using decimal files instead of hex files.
        coincidence_list = scan_times(master1, selected_window)
##            print("Writing coincidences to file:", outputFileName)
        os.chdir(newDirectory)
        fp = open(outputFileName, "a")
        for entry in coincidence_list:
            for subentries in entry:
                print(subentries, file=fp, end=", ")
            print("", file=fp)
        fp.close()
        print("Finished. There were", len(coincidence_list), "coincidences.")
        print("The length of the master list was:", len(master1))           
        os.chdir(folder)

    # ANTI COINCIDENCE
    # Delete data from the telescopes that are anti-coincident.
    # If just one telescope from the perimeter lit up,
    # then do not include that data.
    layout = [[sg.Text("Should we do anti-coincidence scanning?")],
              [sg.Radio("Yes", "RADIO1", default=True), sg.Radio("No", "RADIO1")],
              [sg.Submit()]] 

    window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
    event, values = window.Read()
    window.Close()

    # For radio GUI objects, the 'values' return array is special (must use manual indexing to get return values)
    # Find out which of the options was selected, and store it into a variable that can be interpreted more easily
    answer = 0
    # Assign a value of 1, 2, or 3 to "answer"
    for x in range(len(values)):
        if values[x] == True:
            answer = x + 1
    
    if (answer == 1):
        os.chdir(folder)
        while True:
            layout = [[sg.Text("How many different telescopes were on the periemter?")],
                      [sg.Text("Note: there is no limit on how many telescopes that can be analyzed.")],
                      [sg.Spin([i for i in range(2, 99)], initial_value=2, change_submits=True, key="spin", size=(15, 1))],
                      [sg.Submit()]]

            window = sg.Window("Analyze Cosmic Ray Telescope Data", layout)
            event, value = window.Read()
            while event is not None and event and event is not "Submit":
                telescopeCount = int(value["spin"])
                window.FindElement("spin").Update(telescopeCount)
                event, value = window.Read()

            try:
                count = int(value["spin"])
            except ValueError:
                layout = [[sg.Text("Error: It doesn't make sense to try typing a character here. ", text_color="red")],
                          [sg.Text("Use base 10 numeric digits only.", text_color="red")],
                          [sg.Ok()]]

                window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
                event, values = window.Read()
                
            else:
                break
        window.Close()
        perimeter = []
        layout = [[sg.Text("Look at these telescope file names:")],
                  [sg.Text("Identify the telescopes on the perimeter by using the checkboxes.")]]

        layout2 = []
        counter = 0
        fileList = []
        for file in glob.glob("*.txt"):
            fileList.append(file)
        for x in range(len(fileList)):
            layout2.append([sg.Checkbox(fileList[x])])

        layout.append([sg.Frame("Perimeter telescopes:", layout2, title_color="blue")])
        layout.append([sg.Submit()])
        window = sg.Window('Analyze Cosmic Ray Telescope Data', layout)
        event, values = window.Read()
        window.Close()

        # Perimeter is now an array of integers corresponding to the selected telescopes.
        for x in range(len(values)):
            if values[x] is True:
                perimeter.append(x)

        print("Creating separate list of coincidences without the perimeter telescopes...")
        anti_list = []
        for z in range(len(coincidence_list)):
            coincident_telescopes = coincidence_list[z][3]
            perimeter_count = 0
            for val in perimeter:
                if val in coincident_telescopes:
                    perimeter_count += 1
                    break 
            if perimeter_count == 0:
                anti_list.append(coincidence_list[z])
                
        anti_coinc = 0
        fp = open(outputFileName[:-4] + " anti.txt", "a")
        for data in anti_list:
            for subentries in data:
                print(subentries, file=fp, end=", ")
            print("", file=fp)
            anti_coinc += 1
        fp.close()
        print("Finished writing anti-coincidences. There were", anti_coinc, "of these.")

# CASE 3: Convert from DEC to HEX
# This block of code won't make much functional sense because it is designed without many conditional statements
# But still works for all different kinds of data input that could be decimal, or only integers, or comma separated.
elif (choice == 3):
    print("We will now attempt to convert a decimal file back into hex format.")
    answer = input("What is the name of this decimal file, without .txt?\n") + ".txt"
    while not os.path.isfile(outputFileName):
        print("That decimal file doesn't exist, or the directory needs to be changed.")
        answer = input("Try again, don't include the file extension.\n")
    print("\nAlso, type 1 if the subseconds are in decimal format.")
    multiply = input("Type 0 if the subseconds are in integer format.\n")
    # Read the data from the given file and start by converting each number to hex
    print("Reading data from", answer)
    with open(answer, "r") as file:          # open file
        data = []                            # declare list
        for entry in file:                   # for each value in the file...
            tempVal = entry.rstrip()
            val1 = str()
            val2 = str()
            i = 0
            char = []
            char = tempVal[i]
            while ((char != ',') and (i < len(tempVal) - 1)):
                val1 += char
                i += 1
                char = tempVal[i]
            i += 2
            while (i < len(tempVal)):
                char = tempVal[i]
                val2 += char
                i += 1
            # We could use int(val, 16) below but we would have to convert to
            # string anyways to add leading 0's
##                print("Val1 and Val2 were:", val1, val2)
            if (multiply == '1'):
                val2 = (float(val1) % 1) * LM555FACTOR
                val1 = int(float((val1)))
            # Use ceil because earlier we used int to convert from hex to decimal
            data.append([hex(int(val1)), hex(int(math.ceil(val2)))])
##                print("Now they are:", val1, val2)
            
    # Slice off the useless '0x' character in front of each hex string
    # Also add leading 0's where appropriate
    print("Formatting each hex string...")
    fixedData = []
    for hexdata in data:
        seconds = hexdata[0]
        seconds = seconds[2:]
        while (len(seconds) < 6):
            seconds = '0' + seconds
        subseconds = hexdata[1]
        subseconds = subseconds[2:]
        while (len(subseconds) < 2):
            subseconds = '0' + subseconds
        fixedData.append([seconds, subseconds])
        
    # Write the fixed data to the output file
    print("Writing fixed hex strings to the new file...")
    fp = open(outputFileName, "a")
    counter = 0
    for entry in fixedData: 
        print(entry[0], file=fp, end="")
        print(entry[1], file=fp)
        counter += 1
    fp.close()
    print("Finished.", counter, "events were converted back to HEX.\n")

# Flashdrives cannot be ejected if Python is inside a flashdrive directory.
# So we are going to move to the default directory to allow ejection.
os.chdir(defaultDirectory)
print("You may now eject a flashdrive, but clear it and copy the files first.")
